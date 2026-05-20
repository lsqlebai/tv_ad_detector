#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import subprocess
import sys
import threading
import time
from dataclasses import dataclass, field
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
INPUT_DIR = ROOT / "input"
OUTPUT_DIR = ROOT / "output"
REVIEW_XLSX = OUTPUT_DIR / "ad_review.xlsx"
CLEANED_DIR = OUTPUT_DIR / "cleaned"
DETECT_DIR = OUTPUT_DIR / "detect"


@dataclass
class JobState:
    running: bool = False
    name: str = ""
    started_at: float = 0.0
    finished_at: float = 0.0
    returncode: int | None = None
    log: list[str] = field(default_factory=list)


JOB = JobState()
JOB_LOCK = threading.Lock()


def rel(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def file_item(path: Path) -> dict:
    stat = path.stat()
    return {
        "name": path.name,
        "path": rel(path),
        "size": stat.st_size,
        "mtime": stat.st_mtime,
    }


def list_files(directory: Path, pattern: str) -> list[dict]:
    if not directory.exists():
        return []
    return [file_item(path) for path in sorted(directory.glob(pattern)) if path.is_file()]


def valid_input_files(values: list[str]) -> list[str]:
    available = {item["name"] for item in list_files(INPUT_DIR, "*.mp4")}
    return [value for value in values if value in available]


def safe_cleaned_file(value: str) -> Path | None:
    name = Path(value).name
    path = (CLEANED_DIR / name).resolve()
    if path.parent != CLEANED_DIR.resolve() or path.suffix.lower() != ".mp4":
        return None
    return path if path.exists() and path.is_file() else None


def snapshot() -> dict:
    with JOB_LOCK:
        job = {
            "running": JOB.running,
            "name": JOB.name,
            "started_at": JOB.started_at,
            "finished_at": JOB.finished_at,
            "returncode": JOB.returncode,
            "log": JOB.log[-300:],
        }
    return {
        "job": job,
        "inputs": list_files(INPUT_DIR, "*.mp4"),
        "review": file_item(REVIEW_XLSX) if REVIEW_XLSX.exists() else None,
        "cleaned": list_files(CLEANED_DIR, "*.mp4"),
    }


def append_log(line: str) -> None:
    with JOB_LOCK:
        JOB.log.append(line.rstrip())
        if len(JOB.log) > 1000:
            del JOB.log[: len(JOB.log) - 1000]


def run_job(name: str, command: list[str]) -> None:
    with JOB_LOCK:
        JOB.running = True
        JOB.name = name
        JOB.started_at = time.time()
        JOB.finished_at = 0.0
        JOB.returncode = None
        JOB.log = [f"$ {' '.join(command)}"]

    try:
        process = subprocess.Popen(
            command,
            cwd=ROOT,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        assert process.stdout is not None
        for line in process.stdout:
            append_log(line)
        returncode = process.wait()
    except Exception as exc:  # pragma: no cover - surfaced in the UI log.
        append_log(f"ERROR: {exc}")
        returncode = 1

    with JOB_LOCK:
        JOB.running = False
        JOB.finished_at = time.time()
        JOB.returncode = returncode
        JOB.log.append(f"Finished with exit code {returncode}")


def start_job(name: str, command: list[str]) -> tuple[bool, str]:
    with JOB_LOCK:
        if JOB.running:
            return False, f"{JOB.name} is already running"
    thread = threading.Thread(target=run_job, args=(name, command), daemon=True)
    thread.start()
    return True, "started"


def safe_output_path(value: str) -> Path | None:
    raw_value = unquote(value)
    requested = (ROOT / raw_value).resolve()
    allowed_roots = [OUTPUT_DIR.resolve(), INPUT_DIR.resolve()]
    if any(requested == root or root in requested.parents for root in allowed_roots):
        return requested
    output_requested = (OUTPUT_DIR / raw_value).resolve()
    if output_requested == OUTPUT_DIR.resolve() or OUTPUT_DIR.resolve() in output_requested.parents:
        return output_requested
    return None


def file_url(path: str, inline: bool = False) -> str:
    suffix = "?inline=1" if inline else ""
    return f"/files/{quote(path)}{suffix}"


def parse_time_value(value: str) -> float:
    parts = [float(part) for part in value.strip().split(":")]
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        minutes, seconds = parts
        return minutes * 60 + seconds
    if len(parts) == 3:
        hours, minutes, seconds = parts
        return hours * 3600 + minutes * 60 + seconds
    raise ValueError(f"Bad timestamp: {value!r}")


def find_review_table(ws) -> tuple[int, dict[str, int]]:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        values = [cell.value for cell in row]
        if "delete" in values and "file" in values and "start" in values and "end" in values:
            headers = {str(cell.value): cell.column for cell in row if cell.value}
            return row[0].row, headers
    raise RuntimeError("Could not find review headers")


def detect_frame_indexes() -> tuple[dict[tuple[str, float, float], dict], dict[str, list[dict]]]:
    by_key: dict[tuple[str, float, float], dict] = {}
    by_file: dict[str, list[dict]] = {}
    if not DETECT_DIR.exists():
        return by_key, by_file
    for csv_path in sorted(DETECT_DIR.glob("*.ads.csv")):
        video_name = f"{csv_path.name[:-len('.ads.csv')]}.mp4"
        with csv_path.open(newline="", encoding="utf-8") as handle:
            for row in csv.DictReader(handle):
                frame_data = {
                    "start_frame": row.get("start_frame", ""),
                    "middle_frame": row.get("middle_frame", ""),
                    "end_frame": row.get("end_frame", ""),
                }
                by_file.setdefault(video_name, []).append(frame_data)
                try:
                    start = round(float(row.get("start_seconds") or 0.0), 3)
                    end = round(float(row.get("end_seconds") or 0.0), 3)
                except ValueError:
                    continue
                by_key[(video_name, start, end)] = frame_data
    return by_key, by_file


def read_review_rows() -> list[dict]:
    if not REVIEW_XLSX.exists():
        return []
    wb = load_workbook(REVIEW_XLSX, data_only=True)
    if "Review" not in wb.sheetnames:
        return []
    ws = wb["Review"]
    header_row, headers = find_review_table(ws)
    frames_by_key, frames_by_file = detect_frame_indexes()
    file_offsets: dict[str, int] = {}
    rows: list[dict] = []
    optional = ["score", "kind", "review_required", "sources", "start_frame", "middle_frame", "end_frame"]
    for row_index in range(header_row + 1, ws.max_row + 1):
        file_name = str(ws.cell(row=row_index, column=headers["file"]).value or "").strip()
        if not file_name:
            continue
        row = {
            "row": row_index,
            "delete": str(ws.cell(row=row_index, column=headers["delete"]).value or "NO").strip().upper(),
            "file": file_name,
            "start": str(ws.cell(row=row_index, column=headers["start"]).value or "").strip(),
            "end": str(ws.cell(row=row_index, column=headers["end"]).value or "").strip(),
        }
        for name in optional:
            row[name] = str(ws.cell(row=row_index, column=headers[name]).value or "").strip() if name in headers else ""
        if not any(row.get(name) for name in ("start_frame", "middle_frame", "end_frame")):
            try:
                start_seconds = round(parse_time_value(row["start"]), 3)
                end_seconds = round(parse_time_value(row["end"]), 3)
            except ValueError:
                start_seconds = end_seconds = -1.0
            frame_data = frames_by_key.get((file_name, start_seconds, end_seconds))
            if frame_data is None:
                offset = file_offsets.get(file_name, 0)
                frame_rows = frames_by_file.get(file_name, [])
                frame_data = frame_rows[offset] if offset < len(frame_rows) else None
                file_offsets[file_name] = offset + 1
            if frame_data:
                row.update(frame_data)
        for name in ("start_frame", "middle_frame", "end_frame"):
            frame_path = row.get(name, "")
            if frame_path:
                row[f"{name}_url"] = file_url(frame_path, inline=True)
                row[f"{name}_download_url"] = file_url(frame_path, inline=False)
        rows.append(row)
    return rows


def save_review_rows(rows: list[dict]) -> int:
    if not REVIEW_XLSX.exists():
        raise RuntimeError("Review workbook does not exist")
    wb = load_workbook(REVIEW_XLSX)
    if "Review" not in wb.sheetnames:
        raise RuntimeError("No Review sheet in workbook")
    ws = wb["Review"]
    _, headers = find_review_table(ws)
    changed = 0
    for item in rows:
        row_index = int(item.get("row") or 0)
        if row_index <= 0 or row_index > ws.max_row:
            continue
        delete_value = str(item.get("delete") or "NO").strip().upper()
        if delete_value not in {"YES", "NO"}:
            raise ValueError(f"Row {row_index}: delete must be YES or NO")
        start_value = str(item.get("start") or "").strip()
        end_value = str(item.get("end") or "").strip()
        if not start_value or not end_value:
            raise ValueError(f"Row {row_index}: start/end cannot be empty")
        start_seconds = parse_time_value(start_value)
        end_seconds = parse_time_value(end_value)
        if end_seconds <= start_seconds:
            raise ValueError(f"Row {row_index}: end must be after start")

        ws.cell(row=row_index, column=headers["delete"], value=delete_value)
        ws.cell(row=row_index, column=headers["start"], value=start_value)
        ws.cell(row=row_index, column=headers["end"], value=end_value)
        if "start_seconds" in headers:
            ws.cell(row=row_index, column=headers["start_seconds"], value=start_seconds)
        if "end_seconds" in headers:
            ws.cell(row=row_index, column=headers["end_seconds"], value=end_seconds)
        changed += 1
    wb.save(REVIEW_XLSX)
    return changed


INDEX_HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>TV Ad Detector</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #15181d;
      --muted: #667085;
      --line: #d8dde6;
      --accent: #1463ff;
      --ok: #137a4f;
      --warn: #a35c00;
      --danger: #b42318;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font: 14px/1.45 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    header {
      border-bottom: 1px solid var(--line);
      background: var(--panel);
      padding: 18px 24px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
    }
    h1 { font-size: 20px; margin: 0; letter-spacing: 0; }
    main {
      width: min(1180px, calc(100vw - 32px));
      margin: 20px auto 40px;
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
    }
    section {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      min-width: 0;
    }
    .wide { grid-column: 1 / -1; }
    h2 { font-size: 15px; margin: 0 0 12px; letter-spacing: 0; }
    .actions { display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }
    button, select, input {
      border: 1px solid var(--line);
      background: #fff;
      color: var(--text);
      border-radius: 6px;
      height: 36px;
      padding: 0 12px;
      font: inherit;
    }
    button {
      background: var(--accent);
      border-color: var(--accent);
      color: white;
      cursor: pointer;
      font-weight: 600;
    }
    button.secondary { background: #fff; color: var(--text); border-color: var(--line); }
    button:disabled { opacity: 0.5; cursor: not-allowed; }
    label { color: var(--muted); display: inline-flex; gap: 6px; align-items: center; }
    input[type="number"] { width: 88px; }
    input.time { width: 86px; padding: 0 8px; }
    table { width: 100%; border-collapse: collapse; }
    th, td {
      border-bottom: 1px solid var(--line);
      padding: 9px 6px;
      text-align: left;
      vertical-align: middle;
      overflow-wrap: normal;
    }
    td.file { max-width: 250px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
    td.compact, th.compact { white-space: nowrap; width: 1%; }
    th { color: var(--muted); font-weight: 600; font-size: 12px; }
    td img {
      display: block;
      width: 116px;
      aspect-ratio: 16 / 9;
      object-fit: cover;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #f0f2f5;
    }
    a { color: var(--accent); text-decoration: none; }
    .muted { color: var(--muted); }
    .pill {
      display: inline-flex;
      align-items: center;
      min-height: 24px;
      border-radius: 999px;
      padding: 2px 9px;
      border: 1px solid var(--line);
      color: var(--muted);
      background: #fff;
      font-size: 12px;
    }
    .pill.ok { color: var(--ok); border-color: #b8decf; background: #eefaf5; }
    .pill.warn { color: var(--warn); border-color: #f1d0a3; background: #fff7ed; }
    .pill.danger { color: var(--danger); border-color: #f0b8b8; background: #fff1f1; }
    .table-wrap { overflow-x: auto; }
    .review-actions {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      margin-bottom: 12px;
    }
    .frame-strip { display: grid; grid-template-columns: repeat(3, 116px); gap: 8px; }
    .module-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 12px;
    }
    .module-head h2 { margin: 0; }
    input[type="checkbox"] { width: 16px; height: 16px; padding: 0; }
    pre {
      margin: 0;
      background: #111827;
      color: #e5e7eb;
      border-radius: 8px;
      padding: 14px;
      min-height: 260px;
      max-height: 420px;
      overflow: auto;
      white-space: pre-wrap;
      word-break: break-word;
      font: 12px/1.45 ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
    }
    @media (max-width: 800px) {
      main { grid-template-columns: 1fr; width: calc(100vw - 20px); }
      header { align-items: flex-start; flex-direction: column; }
    }
  </style>
</head>
<body>
  <header>
    <div>
      <h1>TV Ad Detector</h1>
      <div class="muted">MP4 广告候选审核与批量裁剪</div>
    </div>
    <span id="state" class="pill">加载中</span>
  </header>
  <main>
    <section>
      <div class="module-head">
        <h2>输入视频</h2>
        <div class="actions">
          <button id="buildBtn" onclick="startBuild()">生成选中视频的审核表</button>
        </div>
      </div>
      <div id="inputs" class="muted">暂无</div>
    </section>
    <section>
      <div class="module-head">
        <h2>裁剪结果</h2>
        <div class="actions">
          <button id="deleteCleanedBtn" class="secondary" onclick="deleteCleaned()">删除选中结果</button>
        </div>
      </div>
      <div id="cleaned" class="muted">暂无</div>
    </section>
    <section class="wide">
      <div class="review-actions">
        <h2 style="margin:0">候选审核</h2>
        <div class="actions">
          <span id="review" class="muted">暂无审核表</span>
          <button id="saveReviewBtn" class="secondary" onclick="saveReview()" disabled>保存审核</button>
          <button id="cutBtn" onclick="startCut()">按当前审核裁剪</button>
          <label>模式
            <select id="mode">
              <option value="reencode">reencode</option>
              <option value="copy">copy</option>
            </select>
          </label>
          <label>padding
            <input id="padding" type="number" min="0" step="0.1" value="0">
          </label>
          <span id="reviewStatus" class="muted"></span>
        </div>
      </div>
      <div id="reviewRows" class="muted">生成审核表后可在这里编辑</div>
    </section>
    <section class="wide">
      <h2>日志</h2>
      <pre id="log"></pre>
    </section>
  </main>
  <script>
    const stateEl = document.getElementById("state");
    const buildBtn = document.getElementById("buildBtn");
    const cutBtn = document.getElementById("cutBtn");
    const deleteCleanedBtn = document.getElementById("deleteCleanedBtn");
    const saveReviewBtn = document.getElementById("saveReviewBtn");
    const reviewRowsEl = document.getElementById("reviewRows");
    const reviewStatusEl = document.getElementById("reviewStatus");
    const logEl = document.getElementById("log");
    let reviewRows = [];
    let reviewDirty = false;
    let selectedInputs = new Set();
    let selectedCleaned = new Set();
    let initializedInputSelection = false;

    function sizeText(bytes) {
      if (bytes > 1024 * 1024 * 1024) return (bytes / 1024 / 1024 / 1024).toFixed(2) + " GB";
      if (bytes > 1024 * 1024) return (bytes / 1024 / 1024).toFixed(1) + " MB";
      if (bytes > 1024) return (bytes / 1024).toFixed(1) + " KB";
      return bytes + " B";
    }
    function fileLink(item) {
      const href = "/files/" + encodeURIComponent(item.path);
      return `<a href="${href}">${escapeHtml(item.name)}</a> <span class="muted">${sizeText(item.size)}</span>`;
    }
    function escapeHtml(value) {
      return String(value).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
    }
    function syncSelection(set, files) {
      const allowed = new Set(files.map(item => item.name));
      for (const value of Array.from(set)) {
        if (!allowed.has(value)) set.delete(value);
      }
    }
    function ensureDefaultInputSelection(files) {
      if (initializedInputSelection) return;
      files.forEach(item => selectedInputs.add(item.name));
      initializedInputSelection = true;
    }
    function setSelection(kind, name, checked) {
      const target = kind === "input" ? selectedInputs : selectedCleaned;
      if (checked) target.add(name);
      else target.delete(name);
      updateSelectionButtons();
    }
    function setSelectionByIndex(kind, index, checked) {
      const files = kind === "input" ? latestInputs : latestCleaned;
      if (!files[index]) return;
      setSelection(kind, files[index].name, checked);
    }
    function setAllSelection(kind, checked) {
      const files = kind === "input" ? latestInputs : latestCleaned;
      const target = kind === "input" ? selectedInputs : selectedCleaned;
      target.clear();
      if (checked) files.forEach(item => target.add(item.name));
      renderSelectableList(kind === "input" ? "inputs" : "cleaned", kind, files);
      updateSelectionButtons();
    }
    function updateSelectionButtons() {
      buildBtn.disabled = currentRunning || selectedInputs.size === 0;
      cutBtn.disabled = currentRunning || !currentHasReview;
      deleteCleanedBtn.disabled = currentRunning || selectedCleaned.size === 0;
    }
    let latestInputs = [];
    let latestCleaned = [];
    let currentRunning = false;
    let currentHasReview = false;
    function renderSelectableList(target, kind, files) {
      const el = document.getElementById(target);
      if (!files.length) {
        el.innerHTML = '<span class="muted">暂无</span>';
        return;
      }
      const selected = kind === "input" ? selectedInputs : selectedCleaned;
      const allChecked = files.length > 0 && files.every(item => selected.has(item.name));
      el.innerHTML = `<table><thead><tr><th class="compact"><input type="checkbox" ${allChecked ? "checked" : ""} onchange="setAllSelection('${kind}', this.checked)"></th><th>文件</th><th class="compact">大小</th></tr></thead><tbody>${files.map((item, index) =>
        `<tr><td class="compact"><input type="checkbox" ${selected.has(item.name) ? "checked" : ""} onchange="setSelectionByIndex('${kind}', ${index}, this.checked)"></td><td>${fileLink(item)}</td><td class="compact">${sizeText(item.size)}</td></tr>`
      ).join("")}</tbody></table>`;
    }
    function setReviewDirty(value) {
      reviewDirty = value;
      saveReviewBtn.disabled = !value;
      reviewStatusEl.textContent = value ? "有未保存修改" : "";
    }
    function updateReviewRow(index, field, value) {
      reviewRows[index][field] = value;
      setReviewDirty(true);
    }
    function renderReviewRows(rows) {
      reviewRows = rows;
      if (!rows.length) {
        reviewRowsEl.innerHTML = '<span class="muted">暂无候选。请先生成审核表。</span>';
        saveReviewBtn.disabled = true;
        return;
      }
      reviewRowsEl.innerHTML = `<div class="table-wrap"><table><thead><tr>
        <th class="compact">删除</th><th>文件</th><th class="compact">开始</th><th class="compact">结束</th><th class="compact">分数</th><th class="compact">来源</th><th class="compact">状态</th><th>截图</th>
      </tr></thead><tbody>${rows.map((row, index) => {
        const frames = ["start_frame", "middle_frame", "end_frame"].map(key => row[`${key}_url`] ? `<a href="${row[`${key}_download_url`]}"><img src="${row[`${key}_url`]}" alt="${key}" loading="lazy"></a>` : "").join("");
        return `<tr>
          <td class="compact"><select onchange="updateReviewRow(${index}, 'delete', this.value)">
            <option value="YES" ${row.delete === "YES" ? "selected" : ""}>YES</option>
            <option value="NO" ${row.delete !== "YES" ? "selected" : ""}>NO</option>
          </select></td>
          <td class="file" title="${escapeHtml(row.file)}">${escapeHtml(row.file)}</td>
          <td class="compact"><input class="time" value="${escapeHtml(row.start)}" onchange="updateReviewRow(${index}, 'start', this.value)"></td>
          <td class="compact"><input class="time" value="${escapeHtml(row.end)}" onchange="updateReviewRow(${index}, 'end', this.value)"></td>
          <td class="compact">${escapeHtml(row.score || "")}</td>
          <td class="compact">${escapeHtml(row.kind || "")}</td>
          <td class="compact">${row.review_required === "yes" ? '<span class="pill warn">需确认</span>' : '<span class="pill ok">可信</span>'}</td>
          <td style="min-width:380px"><div class="frame-strip">${frames}</div></td>
        </tr>`;
      }).join("")}</tbody></table></div>`;
      setReviewDirty(false);
    }
    async function loadReviewRows(force) {
      if (reviewDirty && !force) return;
      const response = await fetch("/api/review");
      if (!response.ok) return;
      const data = await response.json();
      renderReviewRows(data.rows || []);
    }
    async function refresh() {
      const data = await fetch("/api/status").then(r => r.json());
      const running = data.job.running;
      currentRunning = running;
      currentHasReview = Boolean(data.review);
      stateEl.className = "pill " + (running ? "warn" : data.job.returncode === 0 ? "ok" : data.job.returncode ? "danger" : "");
      stateEl.textContent = running ? `${data.job.name} 运行中` : data.job.returncode === 0 ? "空闲，上次成功" : data.job.returncode ? "空闲，上次失败" : "空闲";
      document.getElementById("review").innerHTML = data.review ? fileLink(data.review) : '<span class="muted">暂无</span>';
      latestInputs = data.inputs;
      latestCleaned = data.cleaned;
      ensureDefaultInputSelection(latestInputs);
      syncSelection(selectedInputs, latestInputs);
      syncSelection(selectedCleaned, latestCleaned);
      renderSelectableList("inputs", "input", latestInputs);
      renderSelectableList("cleaned", "cleaned", latestCleaned);
      updateSelectionButtons();
      logEl.textContent = data.job.log.join("\n");
      logEl.scrollTop = logEl.scrollHeight;
      if (data.review) await loadReviewRows(false);
      else renderReviewRows([]);
    }
    async function postJson(url, payload) {
      const response = await fetch(url, {method: "POST", headers: {"Content-Type": "application/json"}, body: JSON.stringify(payload || {})});
      if (!response.ok) alert(await response.text());
      await refresh();
    }
    async function saveReview() {
      const response = await fetch("/api/review", {method: "POST", headers: {"Content-Type": "application/json"}, body: JSON.stringify({rows: reviewRows})});
      const data = await response.json().catch(() => ({}));
      if (!response.ok) {
        alert(data.message || "保存失败");
        return;
      }
      setReviewDirty(false);
      reviewStatusEl.textContent = `已保存 ${data.changed} 行`;
      await loadReviewRows(true);
    }
    function startBuild() {
      if (selectedInputs.size === 0) {
        alert("请先选择要处理的输入视频。");
        return;
      }
      postJson("/api/build-review", {files: Array.from(selectedInputs)});
    }
    function startCut() {
      if (reviewDirty) {
        alert("请先保存审核修改，再执行裁剪。");
        return;
      }
      postJson("/api/cut", {
        mode: document.getElementById("mode").value,
        padding: document.getElementById("padding").value
      });
    }
    function deleteCleaned() {
      if (selectedCleaned.size === 0) {
        alert("请先选择要删除的裁剪结果。");
        return;
      }
      if (!confirm(`删除 ${selectedCleaned.size} 个裁剪结果？`)) return;
      postJson("/api/delete-cleaned", {files: Array.from(selectedCleaned)});
    }
    refresh();
    setInterval(refresh, 2500);
  </script>
</body>
</html>
"""


class Handler(BaseHTTPRequestHandler):
    server_version = "TVAdDetector/1.0"

    def log_message(self, fmt: str, *args: object) -> None:
        return

    def send_bytes(self, body: bytes, content_type: str, status: int = 200) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_json(self, payload: dict, status: int = 200) -> None:
        self.send_bytes(json.dumps(payload, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8", status)

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        body = self.rfile.read(length).decode("utf-8")
        if self.headers.get("Content-Type", "").startswith("application/json"):
            return json.loads(body or "{}")
        values = parse_qs(body)
        return {key: value[-1] for key, value in values.items()}

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self.send_bytes(INDEX_HTML.encode("utf-8"), "text/html; charset=utf-8")
            return
        if parsed.path == "/api/status":
            self.send_json(snapshot())
            return
        if parsed.path == "/api/review":
            try:
                self.send_json({"rows": read_review_rows()})
            except Exception as exc:
                self.send_json({"rows": [], "message": str(exc)}, 500)
            return
        if parsed.path.startswith("/files/"):
            raw_file_path = unquote(parsed.path[len("/files/") :])
            path = safe_output_path(raw_file_path)
            if path is None or not path.is_file():
                self.send_error(HTTPStatus.NOT_FOUND, "File not found")
                return
            content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(path.stat().st_size))
            disposition = "inline" if parse_qs(parsed.query).get("inline") else "attachment"
            self.send_header("Content-Disposition", f"{disposition}; filename*=UTF-8''{quote(path.name)}")
            self.end_headers()
            with path.open("rb") as handle:
                while chunk := handle.read(1024 * 1024):
                    self.wfile.write(chunk)
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        payload = self.read_json()
        if parsed.path == "/api/build-review":
            files = valid_input_files(list(payload.get("files") or []))
            if not files:
                self.send_json({"ok": False, "message": "请选择要处理的输入视频"}, 400)
                return
            command = [sys.executable, str(ROOT / "scripts" / "build_review.py"), "--files", *files]
            ok, message = start_job("生成审核表", command)
            self.send_json({"ok": ok, "message": message}, 200 if ok else 409)
            return
        if parsed.path == "/api/cut":
            mode = str(payload.get("mode") or "reencode")
            if mode not in {"copy", "reencode"}:
                self.send_json({"ok": False, "message": "bad mode"}, 400)
                return
            try:
                padding = max(0.0, float(payload.get("padding") or 0.0))
            except ValueError:
                self.send_json({"ok": False, "message": "bad padding"}, 400)
                return
            command = [
                sys.executable,
                str(ROOT / "scripts" / "cut_ads.py"),
                "--mode",
                mode,
                "--padding",
                f"{padding:.3f}",
            ]
            ok, message = start_job("裁剪视频", command)
            self.send_json({"ok": ok, "message": message}, 200 if ok else 409)
            return
        if parsed.path == "/api/delete-cleaned":
            deleted = []
            for value in list(payload.get("files") or []):
                path = safe_cleaned_file(str(value))
                if path is None:
                    continue
                path.unlink()
                deleted.append(path.name)
            self.send_json({"ok": True, "deleted": deleted})
            return
        if parsed.path == "/api/review":
            try:
                changed = save_review_rows(payload.get("rows") or [])
            except Exception as exc:
                self.send_json({"ok": False, "message": str(exc)}, 400)
                return
            self.send_json({"ok": True, "changed": changed})
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Not found")


def main() -> int:
    parser = argparse.ArgumentParser(description="Run the TV ad detector web console.")
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", type=int, default=8787)
    args = parser.parse_args()

    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CLEANED_DIR.mkdir(parents=True, exist_ok=True)

    server = ThreadingHTTPServer((args.host, args.port), Handler)
    print(f"TV Ad Detector web console: http://{args.host}:{args.port}", flush=True)
    server.serve_forever()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
