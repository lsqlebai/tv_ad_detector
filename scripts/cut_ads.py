#!/usr/bin/env python3
import argparse
import re
import shlex
import subprocess
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Optional

import imageio_ffmpeg
from openpyxl import load_workbook

def parse_time(value: str) -> float:
    parts = [float(part) for part in value.strip().split(":")]
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        minutes, seconds = parts
        return minutes * 60 + seconds
    if len(parts) == 3:
        hours, minutes, seconds = parts
        return hours * 3600 + minutes * 60 + seconds
    raise ValueError(f"Bad timestamp: {value!r}")


def format_time(seconds: float) -> str:
    seconds = max(0.0, seconds)
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    remaining = seconds - hours * 3600 - minutes * 60
    return f"{hours:02d}:{minutes:02d}:{remaining:06.3f}"


def format_range(start: float, end: float) -> str:
    return f"{format_time(start)}-{format_time(end)} ({end - start:.3f}s)"


def probe_duration(ffmpeg: str, video_path: Path) -> float:
    result = subprocess.run(
        [ffmpeg, "-hide_banner", "-i", str(video_path)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        check=False,
    )
    output = result.stderr + result.stdout
    match = re.search(r"Duration:\s*(\d+):(\d+):(\d+(?:\.\d+)?)", output)
    if not match:
        _, seconds = imageio_ffmpeg.count_frames_and_secs(str(video_path))
        duration = float(seconds)
    else:
        hours, minutes, seconds = match.groups()
        duration = int(hours) * 3600 + int(minutes) * 60 + float(seconds)
    return duration


def read_ad_ranges(path: Path, duration: float) -> list[tuple[float, float]]:
    ranges: list[tuple[float, float]] = []
    if not path.exists():
        return ranges

    for line in path.read_text(encoding="utf-8").splitlines():
        value = line.strip()
        if not value or value.startswith("#"):
            continue
        if "-" not in value:
            continue
        start_raw, end_raw = value.split("-", 1)
        start = parse_time(start_raw)
        end = parse_time(end_raw) if end_raw.strip() else duration
        start = max(0.0, min(duration, start))
        end = max(0.0, min(duration, end))
        if end > start:
            ranges.append((start, end))

    ranges.sort()
    merged: list[tuple[float, float]] = []
    for start, end in ranges:
        if not merged or start > merged[-1][1]:
            merged.append((start, end))
        else:
            merged[-1] = (merged[-1][0], max(merged[-1][1], end))
    return merged


def merge_ranges(ranges: list[tuple[float, float]]) -> list[tuple[float, float]]:
    ranges = sorted(ranges)
    merged: list[tuple[float, float]] = []
    for start, end in ranges:
        if end <= start:
            continue
        if not merged or start > merged[-1][1]:
            merged.append((start, end))
        else:
            merged[-1] = (merged[-1][0], max(merged[-1][1], end))
    return merged


def read_review_ranges(path: Path) -> dict[str, list[tuple[float, float]]]:
    wb = load_workbook(path, data_only=True)
    if "Review" not in wb.sheetnames:
        raise RuntimeError(f"No Review sheet in {path}")
    ws = wb["Review"]

    header_row = None
    headers: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        values = [cell.value for cell in row]
        if "delete" in values and "file" in values and "start" in values and "end" in values:
            header_row = row[0].row
            headers = {str(cell.value): cell.column for cell in row if cell.value}
            break
    if header_row is None:
        raise RuntimeError(f"Could not find review headers in {path}")

    required = ["delete", "file", "start", "end"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f"Missing columns in {path}: {', '.join(missing)}")

    ranges_by_file: dict[str, list[tuple[float, float]]] = {}
    for row_index in range(header_row + 1, ws.max_row + 1):
        decision = ws.cell(row=row_index, column=headers["delete"]).value
        if str(decision or "").strip().upper() not in {"YES", "Y", "TRUE", "1"}:
            continue
        file_name = str(ws.cell(row=row_index, column=headers["file"]).value or "").strip()
        start_value = ws.cell(row=row_index, column=headers["start"]).value
        end_value = ws.cell(row=row_index, column=headers["end"]).value
        end_after_value = ws.cell(row=row_index, column=headers["end_after_seconds"]).value if "end_after_seconds" in headers else None
        if start_value in (None, "") and "start_seconds" in headers:
            start_value = ws.cell(row=row_index, column=headers["start_seconds"]).value
        if end_value in (None, "") and "end_seconds" in headers:
            end_value = ws.cell(row=row_index, column=headers["end_seconds"]).value
        if not file_name or start_value in (None, "") or end_value in (None, ""):
            continue
        start = parse_time(str(start_value))
        end = parse_time(str(end_after_value)) if end_after_value not in (None, "") else parse_time(str(end_value))
        ranges_by_file.setdefault(file_name, []).append((start, end))

    return {file_name: merge_ranges(ranges) for file_name, ranges in ranges_by_file.items()}


def keep_ranges(ad_ranges: list[tuple[float, float]], duration: float, padding: float) -> list[tuple[float, float]]:
    adjusted = []
    for start, end in ad_ranges:
        adjusted.append((max(0.0, start - padding), min(duration, end + padding)))

    keeps: list[tuple[float, float]] = []
    cursor = 0.0
    for start, end in adjusted:
        if start > cursor:
            keeps.append((cursor, start))
        cursor = max(cursor, end)
    if cursor < duration:
        keeps.append((cursor, duration))
    return [(start, end) for start, end in keeps if end - start >= 0.2]


def run(command: list[str]) -> None:
    print(f"  $ {shlex.join(command)}", flush=True)
    subprocess.run(command, check=True)


def cut_segment(ffmpeg: str, source: Path, target: Path, start: float, end: float, mode: str) -> None:
    if mode == "copy":
        command = [
            ffmpeg,
            "-hide_banner",
            "-loglevel",
            "error",
            "-y",
            "-ss",
            format_time(start),
            "-i",
            str(source),
            "-t",
            format_time(end - start),
        ]
        command += ["-c", "copy", "-avoid_negative_ts", "make_zero"]
    else:
        command = [
            ffmpeg,
            "-hide_banner",
            "-loglevel",
            "error",
            "-y",
            "-i",
            str(source),
            "-ss",
            format_time(start),
            "-to",
            format_time(end),
            "-c:v",
            "libx264",
            "-preset",
            "veryfast",
            "-crf",
            "18",
            "-c:a",
            "aac",
            "-b:a",
            "160k",
        ]
    command.append(str(target))
    run(command)


def concat_segments(ffmpeg: str, segments: list[Path], output_path: Path) -> None:
    list_path = output_path.with_suffix(".concat.txt")
    lines = []
    for segment in segments:
        escaped = str(segment.resolve()).replace("'", "'\\''")
        lines.append(f"file '{escaped}'")
    list_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    try:
        run(
            [
                ffmpeg,
                "-hide_banner",
                "-loglevel",
                "error",
                "-y",
                "-f",
                "concat",
                "-safe",
                "0",
                "-i",
                str(list_path.resolve()),
                "-c",
                "copy",
                str(output_path.resolve()),
            ]
        )
    finally:
        list_path.unlink(missing_ok=True)


def process_video(
    video_path: Path,
    ads_dir: Path,
    output_dir: Path,
    mode: str,
    padding: float,
    review_ranges: Optional[dict[str, list[tuple[float, float]]]] = None,
) -> None:
    ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
    started = time.perf_counter()
    duration = probe_duration(ffmpeg, video_path)
    if review_ranges is None:
        ad_path = ads_dir / f"{video_path.stem}.ads.txt"
        ads = read_ad_ranges(ad_path, duration)
        source = str(ad_path)
    else:
        ads = review_ranges.get(video_path.name, [])
        source = "review workbook"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / video_path.name

    print(f"{video_path.name}: duration={format_time(duration)} mode={mode} padding={padding:.3f}s source={source}", flush=True)
    if not ads:
        print(f"{video_path.name}: no ads in {source}, skipping")
        return

    keeps = keep_ranges(ads, duration, padding)
    if not keeps:
        print(f"{video_path.name}: all content marked as ads, skipping")
        return

    print(f"{video_path.name}: remove {len(ads)} range(s)", flush=True)
    for index, (start, end) in enumerate(ads, start=1):
        print(f"  ad {index:02d}: {format_range(start, end)}", flush=True)
    print(f"{video_path.name}: keep {len(keeps)} segment(s)", flush=True)
    for index, (start, end) in enumerate(keeps, start=1):
        print(f"  keep {index:02d}: {format_range(start, end)}", flush=True)

    tmp_dir = output_dir / f".{video_path.stem}.segments"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    segments: list[Path] = []
    try:
        jobs = min(2, len(keeps)) if mode == "reencode" else 1
        segment_specs = []
        for index, (start, end) in enumerate(keeps, start=1):
            segment_path = tmp_dir / f"{index:03d}.mp4"
            segment_specs.append((index, start, end, segment_path))
            segments.append(segment_path)
        if jobs > 1:
            print(f"{video_path.name}: cutting {len(keeps)} segment(s) with {jobs} parallel job(s)", flush=True)
            with ThreadPoolExecutor(max_workers=jobs) as executor:
                futures = {}
                for index, start, end, segment_path in segment_specs:
                    print(f"{video_path.name}: cutting segment {index}/{len(keeps)} -> {segment_path.name}", flush=True)
                    future = executor.submit(cut_segment, ffmpeg, video_path, segment_path, start, end, mode)
                    futures[future] = index
                for future in as_completed(futures):
                    future.result()
        else:
            for index, start, end, segment_path in segment_specs:
                print(f"{video_path.name}: cutting segment {index}/{len(keeps)} -> {segment_path.name}", flush=True)
                cut_segment(ffmpeg, video_path, segment_path, start, end, mode)
        print(f"{video_path.name}: concatenating {len(segments)} segment(s) -> {output_path}", flush=True)
        concat_segments(ffmpeg, segments, output_path)
    finally:
        for segment in segments:
            segment.unlink(missing_ok=True)
        try:
            tmp_dir.rmdir()
        except OSError:
            pass

    if output_path.exists():
        print(f"{video_path.name}: segmented cut completed in {time.perf_counter() - started:.1f}s", flush=True)
        print(f"{video_path.name}: removed {len(ads)} ad range(s), wrote {output_path} ({output_path.stat().st_size / 1024 / 1024:.1f} MB)", flush=True)
    else:
        print(f"{video_path.name}: removed {len(ads)} ad range(s), but output file is missing: {output_path}", flush=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="Cut trusted ad ranges from videos using *.ads.txt.")
    parser.add_argument("--input-dir", type=Path, default=Path("input"))
    parser.add_argument("--ads-dir", type=Path, default=Path("output"))
    parser.add_argument("--review-xlsx", type=Path, default=Path("output/ad_review.xlsx"), help="Read delete=YES rows from an ad review workbook.")
    parser.add_argument("--output-dir", type=Path, default=Path("output/cleaned"))
    parser.add_argument("--mode", choices=["copy", "reencode"], default="reencode")
    parser.add_argument("--padding", type=float, default=0.0, help="Seconds to remove before/after each ad range.")
    parser.add_argument("--files", nargs="*", default=[], help="Optional input video filenames to process.")
    args = parser.parse_args()

    videos = sorted(args.input_dir.glob("*.mp4"))
    if args.files:
        wanted = set(args.files)
        videos = [path for path in videos if path.name in wanted]
    if not videos:
        raise SystemExit(f"No .mp4 files found in {args.input_dir}")

    review_ranges = read_review_ranges(args.review_xlsx) if args.review_xlsx else None
    for video_path in videos:
        process_video(video_path, args.ads_dir, args.output_dir, args.mode, args.padding, review_ranges)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
