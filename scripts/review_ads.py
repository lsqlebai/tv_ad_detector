#!/usr/bin/env python3
import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook


HEADERS = [
    "candidate_id",
    "delete",
    "file",
    "start",
    "end",
    "start_seconds",
    "end_seconds",
    "score",
    "kind",
    "review_required",
    "sources",
    "boundary_debug",
    "start_before_frame",
    "start_frame",
    "middle_frame",
    "end_frame",
    "end_after_frame",
]

FRAME_COLUMNS = [
    ("start_before_frame", "M"),
    ("start_frame", "N"),
    ("middle_frame", "O"),
    ("end_frame", "P"),
    ("end_after_frame", "Q"),
]


def format_time(seconds: float) -> str:
    seconds = max(0.0, seconds)
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    remaining = seconds - hours * 3600 - minutes * 60
    if hours:
        return f"{hours}:{minutes:02d}:{remaining:06.3f}"
    return f"{minutes}:{remaining:06.3f}"


def parse_time(value: str) -> float:
    parts = [float(part) for part in value.strip().split(":")]
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        minutes, seconds = parts
        return minutes * 60 + seconds
    if len(parts) == 3:
        hours, minutes, seconds = parts
        return hours * 3600 + minutes * 60 + seconds
    raise ValueError(f"Bad timestamp: {value!r}")


def display_time(row: dict, name: str) -> str:
    seconds = row.get(f"{name}_seconds")
    if seconds not in (None, ""):
        return format_time(float(seconds))
    return row.get(name, "")


def read_candidates(output_dir: Path, files: list[str] | None = None) -> list[dict]:
    wanted = set(files or [])
    rows: list[dict] = []
    for csv_path in sorted(output_dir.glob("*.ads.csv")):
        video_stem = csv_path.name[: -len(".ads.csv")]
        file_name = f"{video_stem}.mp4"
        if wanted and file_name not in wanted:
            continue
        with csv_path.open(newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for index, row in enumerate(reader, start=1):
                row["file"] = file_name
                row["candidate_id"] = f"{video_stem}:{index:03d}"
                row["delete"] = "NO" if row.get("review_required") == "yes" else "YES"
                rows.append(row)
    return rows


def read_existing_reviews(xlsx_path: Path) -> tuple[dict[str, dict], dict[tuple[str, float, float], dict]]:
    if not xlsx_path.exists():
        return {}, {}
    wb = load_workbook(xlsx_path, data_only=True)
    if "Review" not in wb.sheetnames:
        return {}, {}
    ws = wb["Review"]
    header_row = None
    headers: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        values = [cell.value for cell in row]
        if "delete" in values and "file" in values and "start" in values and "end" in values:
            header_row = row[0].row
            headers = {str(cell.value): cell.column for cell in row if cell.value}
            break
    if header_row is None:
        return {}, {}

    by_id: dict[str, dict] = {}
    by_range: dict[tuple[str, float, float], dict] = {}
    for row_index in range(header_row + 1, ws.max_row + 1):
        file_name = str(ws.cell(row=row_index, column=headers["file"]).value or "").strip()
        if not file_name:
            continue
        delete_value = str(ws.cell(row=row_index, column=headers["delete"]).value or "").strip()
        start_value = str(ws.cell(row=row_index, column=headers["start"]).value or "").strip()
        end_value = str(ws.cell(row=row_index, column=headers["end"]).value or "").strip()
        existing = {"delete": delete_value, "start": start_value, "end": end_value}
        candidate_id = str(ws.cell(row=row_index, column=headers["candidate_id"]).value or "").strip() if "candidate_id" in headers else ""
        if candidate_id:
            by_id[candidate_id] = existing
        if "start_seconds" in headers and "end_seconds" in headers:
            start_seconds = ws.cell(row=row_index, column=headers["start_seconds"]).value
            end_seconds = ws.cell(row=row_index, column=headers["end_seconds"]).value
            if start_seconds not in (None, "") and end_seconds not in (None, ""):
                by_range[(file_name, round(float(start_seconds), 3), round(float(end_seconds), 3))] = existing
    return by_id, by_range


def apply_existing_reviews(rows: list[dict], xlsx_path: Path) -> None:
    by_id, by_range = read_existing_reviews(xlsx_path)
    if not by_id and not by_range:
        return
    for row in rows:
        existing = by_id.get(row.get("candidate_id", ""))
        if existing is None:
            key = (
                row.get("file", ""),
                round(float(row["start_seconds"]), 3) if row.get("start_seconds") else -1.0,
                round(float(row["end_seconds"]), 3) if row.get("end_seconds") else -1.0,
            )
            existing = by_range.get(key)
        if not existing:
            continue
        if existing.get("delete"):
            row["delete"] = existing["delete"]
        if existing.get("start"):
            row["start"] = existing["start"]
            row["start_seconds"] = parse_time(existing["start"])
        if existing.get("end"):
            row["end"] = existing["end"]
            row["end_seconds"] = parse_time(existing["end"])


def resolve_image_path(workbook_dir: Path, value: str) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return workbook_dir / path


def add_image_or_label(ws, workbook_dir: Path, value: str, cell: str, width: int, height: int) -> None:
    if value == "__VIDEO_END__":
        target = ws[cell]
        target.value = "视频结束"
        target.alignment = Alignment(horizontal="center", vertical="center")
        target.font = Font(color="666666")
        return
    image_path = resolve_image_path(workbook_dir, value)
    if not image_path.is_file():
        return
    image = Image(str(image_path))
    image.width = width
    image.height = height
    ws.add_image(image, cell)


def row_values(row: dict) -> list:
    return [
        row.get("candidate_id", ""),
        row.get("delete", "NO"),
        row.get("file", ""),
        display_time(row, "start"),
        display_time(row, "end"),
        float(row["start_seconds"]) if row.get("start_seconds") else None,
        float(row["end_seconds"]) if row.get("end_seconds") else None,
        float(row["score"]) if row.get("score") else None,
        row.get("kind", ""),
        row.get("review_required", ""),
        row.get("sources", ""),
        row.get("boundary_debug", ""),
        "",
        "",
        "",
        "",
        "",
    ]


def build_workbook(rows: list[dict], output_dir: Path, xlsx_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    trusted_fill = PatternFill("solid", fgColor="E2F0D9")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    border = Border(bottom=Side(style="thin", color="B7B7B7"))

    ws["A1"] = "Ad Candidate Review"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A2"] = "Set delete=YES for ranges to remove. Rows with review_required=YES are heuristic candidates and need manual confirmation."
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    for column, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    ws.add_data_validation(dv)

    for row_index, row in enumerate(rows, start=5):
        write_row(ws, row_index, row, output_dir, border)
        dv.add(ws.cell(row=row_index, column=2))

    last_row = max(5, len(rows) + 4)
    ws.auto_filter.ref = f"A4:Q{last_row}"
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 42
    ws.column_dimensions["L"].width = 48
    ws.column_dimensions["M"].width = 26
    ws.column_dimensions["N"].width = 26
    ws.column_dimensions["O"].width = 26
    ws.column_dimensions["P"].width = 26
    ws.column_dimensions["Q"].width = 26
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["G"].hidden = True

    ws.conditional_formatting.add(
        f"A5:Q{last_row}",
        FormulaRule(formula=["$B5=\"YES\""], fill=trusted_fill),
    )
    ws.conditional_formatting.add(
        f"A5:Q{last_row}",
        FormulaRule(formula=["$J5=\"yes\""], fill=review_fill),
    )

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A2"] = "Candidate rows"
    summary["B2"] = len(rows)
    summary["A3"] = "Default delete=YES"
    summary["B3"] = sum(1 for row in rows if row.get("delete") == "YES")
    summary["A4"] = "Needs review"
    summary["B4"] = sum(1 for row in rows if row.get("review_required") == "yes")
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 16

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = xlsx_path.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)
    tmp_path.replace(xlsx_path)


def write_row(ws, row_index: int, row: dict, output_dir: Path, border: Border) -> None:
    for column, value in enumerate(row_values(row), start=1):
        cell = ws.cell(row=row_index, column=column, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
    for frame_name, column_letter in FRAME_COLUMNS:
        add_image_or_label(ws, output_dir.parent, row.get(frame_name, ""), f"{column_letter}{row_index}", 180, 101)
    ws.row_dimensions[row_index].height = 84


def workbook_headers(ws) -> tuple[int, dict[str, int]]:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        values = [cell.value for cell in row]
        if "candidate_id" in values and "delete" in values and "file" in values:
            return row[0].row, {str(cell.value): cell.column for cell in row if cell.value}
    return 0, {}


def append_workbook(rows: list[dict], output_dir: Path, xlsx_path: Path) -> int:
    if not xlsx_path.exists():
        build_workbook(rows, output_dir, xlsx_path)
        return len(rows)

    wb = load_workbook(xlsx_path)
    if "Review" not in wb.sheetnames:
        build_workbook(rows, output_dir, xlsx_path)
        return len(rows)
    ws = wb["Review"]
    _, headers = workbook_headers(ws)
    if "candidate_id" not in headers:
        build_workbook(rows, output_dir, xlsx_path)
        return len(rows)

    existing_ids = {
        str(ws.cell(row=row_index, column=headers["candidate_id"]).value or "").strip()
        for row_index in range(5, ws.max_row + 1)
    }
    new_rows = [row for row in rows if row.get("candidate_id", "") not in existing_ids]
    if not new_rows:
        return 0

    border = Border(bottom=Side(style="thin", color="B7B7B7"))
    trusted_fill = PatternFill("solid", fgColor="E2F0D9")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    ws.add_data_validation(dv)
    start_row = ws.max_row + 1
    for offset, row in enumerate(new_rows):
        row_index = start_row + offset
        write_row(ws, row_index, row, output_dir, border)
        dv.add(ws.cell(row=row_index, column=2))

    last_row = ws.max_row
    ws.auto_filter.ref = f"A4:Q{last_row}"
    ws.conditional_formatting.add(
        f"A{start_row}:Q{last_row}",
        FormulaRule(formula=[f"$B{start_row}=\"YES\""], fill=trusted_fill),
    )
    ws.conditional_formatting.add(
        f"A{start_row}:Q{last_row}",
        FormulaRule(formula=[f"$J{start_row}=\"yes\""], fill=review_fill),
    )
    if "Summary" in wb.sheetnames:
        summary = wb["Summary"]
        summary["B2"] = last_row - 4
        summary["B3"] = sum(1 for row_index in range(5, last_row + 1) if ws.cell(row=row_index, column=2).value == "YES")
        summary["B4"] = sum(1 for row_index in range(5, last_row + 1) if ws.cell(row=row_index, column=10).value == "yes")

    tmp_path = xlsx_path.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)
    tmp_path.replace(xlsx_path)
    return len(new_rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Create an Excel workbook for reviewing ad candidates.")
    parser.add_argument("--output-dir", type=Path, default=Path("output/detect"))
    parser.add_argument("--xlsx", type=Path, default=Path("output/ad_review.xlsx"))
    parser.add_argument("--append", action="store_true", help="Append new candidates to an existing workbook instead of rebuilding it.")
    parser.add_argument("--files", nargs="*", default=[], help="Optional video filenames whose candidates should be written.")
    args = parser.parse_args()

    rows = read_candidates(args.output_dir, args.files)
    if not rows:
        raise SystemExit(f"No *.ads.csv files found in {args.output_dir}")
    if args.append:
        count = append_workbook(rows, args.output_dir, args.xlsx)
        print(f"Appended {count} candidate row(s) to review workbook: {args.xlsx}")
    else:
        apply_existing_reviews(rows, args.xlsx)
        build_workbook(rows, args.output_dir, args.xlsx)
        print(f"Wrote review workbook: {args.xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
